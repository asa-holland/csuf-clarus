"""Batch pipeline runner — processes all preprocessed corpus files and writes Excel results."""

import sys
import os
import json
import math
from pathlib import Path
from datetime import datetime

# Make src importable — works both locally (src/clarus) and in Docker (/app/clarus)
_src = Path(__file__).parent / "src"
if _src.exists():
    sys.path.insert(0, str(_src))

import numpy as np

# ── pipeline imports ──────────────────────────────────────────────────────────

print("Loading pipeline modules...")
from clarus.analysis.document_processor import DocumentProcessor
from clarus.analysis.semantic_extractor import SemanticExtractor
from clarus.analysis.terminology_validator import TerminologyValidator
from clarus.analysis.contradiction_analyzer import ContradictionDetector
from clarus.analysis.s2_analyzer import S2Analyzer

import nltk
try:
    nltk.data.find("tokenizers/punkt")
except LookupError:
    nltk.download("punkt")

print("Pipeline modules loaded.\n")

# ── helpers ───────────────────────────────────────────────────────────────────

def safe_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return None

# ── locate corpus files ───────────────────────────────────────────────────────

PREPROCESSED_DIR = Path(__file__).parent / "corpus_files_preprocessed"
OUTPUT_DIR = Path(__file__).parent / "uploads"
OUTPUT_DIR.mkdir(exist_ok=True)

content_files = sorted(PREPROCESSED_DIR.glob("*_content.txt"))
print(f"Found {len(content_files)} documents to process.\n")

# ── run pipeline ──────────────────────────────────────────────────────────────

results = []

for i, content_path in enumerate(content_files, 1):
    doc_name = content_path.stem.replace("_content", "")
    metadata_path = PREPROCESSED_DIR / f"{doc_name}_metadata.json"

    print(f"[{i:02d}/{len(content_files)}] Processing: {doc_name}")

    try:
        text = content_path.read_text(encoding="utf-8", errors="replace")

        meta = {}
        if metadata_path.exists():
            with open(metadata_path, encoding="utf-8") as f:
                meta = json.load(f)

        # 1. Document segmentation
        processor = DocumentProcessor()
        proc_result = processor.process_document(text, metadata=meta)
        stats = proc_result.processing_stats

        # 2. Semantic extraction
        extractor = SemanticExtractor()
        profiles = []
        total_conf = 0.0
        for seg in proc_result.segments:
            profile = extractor.extract_semantic_profile(seg.text)
            required = (profile.modality, profile.subject, profile.object, profile.condition)
            if all(required):
                pdata = {
                    "sentence": profile.original_sentence,
                    "confidence": safe_float(profile.confidence),
                    "modality": profile.modality.text if profile.modality else None,
                    "condition": profile.condition.text if profile.condition else None,
                    "subject": profile.subject.text if profile.subject else None,
                    "object": profile.object.text if profile.object else None,
                    "temporal": profile.temporal.text if profile.temporal else None,
                    "negation": profile.negation.text if profile.negation else None,
                }
                profiles.append(pdata)
                total_conf += safe_float(profile.confidence) or 0.0
        avg_profile_conf = total_conf / len(profiles) if profiles else 0.0

        # 3. Terminology (FAIL-001)
        tv = TerminologyValidator()
        extracted_terms = tv.extract_terms_from_text(text)

        min_quality = 0.4
        max_terms = 50
        high_quality = []
        for term in extracted_terms:
            res = tv.validate_term(term)
            if (
                res.classification.confidence >= min_quality
                and res.classification.is_domain_term
                and not res.classification.is_common_english
            ):
                high_quality.append(term)
            if len(high_quality) >= max_terms:
                break

        if not high_quality:
            high_quality = [
                t for t in extracted_terms
                if " " in t or (t[0].isupper() and not t.isupper())
            ][:max_terms]

        term_result = tv.validate_terms(high_quality, full_text=text)
        term_stats = term_result.statistics

        undefined_terms = [
            {
                "term": r.term,
                "is_acronym": r.term.isupper() and len(r.term) >= 2,
                "context": r.context_excerpt,
                "suggested_def": r.suggested_definition,
            }
            for r in term_result.undefined_terms
        ]

        # 4. Contradiction detection (FAIL-003 / FAIL-004 via NLI)
        candidate_stmts = [
            seg.text for seg in proc_result.segments
            if seg.element_type.value in ("normative", "unknown")
        ]
        detector = ContradictionDetector()
        contradictions = detector.detect_contradictions(candidate_stmts)

        contra_list = [
            {
                "type": c.contradiction_type.value,
                "confidence": safe_float(c.confidence),
                "statement_a": c.statement_a,
                "statement_b": c.statement_b,
                "explanation": c.explanation,
            }
            for c in contradictions
        ]

        # 5. S2 Taxonomy analysis (FAIL-002, 004, 005, 006, 007, 008)
        s2_findings_raw = S2Analyzer().analyze(text, proc_result.segments)
        s2_findings = [
            {
                "uid": f.uid,
                "error_type": f.error_type,
                "category": f.category,
                "severity": f.severity,
                "text_span": f.text_span,
                "context": f.context,
                "explanation": f.explanation,
                "segment_type": f.segment_type,
                "confidence": f.confidence,
            }
            for f in s2_findings_raw
        ]

        # Also wrap FAIL-001 (undefined terms) and FAIL-003 (contradictions)
        # into S2Finding format for the unified sheet
        for ut in undefined_terms:
            s2_findings.append({
                "uid": "FAIL-001",
                "error_type": "Undefined Term",
                "category": "Terminology",
                "severity": "High",
                "text_span": ut["term"],
                "context": (ut["context"] or "")[:300],
                "explanation": f'Term "{ut["term"]}" is used but has no definition in this document.',
                "segment_type": "normative",
                "confidence": 0.80,
            })

        for c in contra_list:
            s2_findings.append({
                "uid": "FAIL-003",
                "error_type": "Direct Contradiction",
                "category": "Logical",
                "severity": "High",
                "text_span": c["statement_a"][:80],
                "context": f'A: {c["statement_a"][:200]} | B: {c["statement_b"][:200]}',
                "explanation": c["explanation"],
                "segment_type": "normative",
                "confidence": safe_float(c["confidence"]) or 0.0,
            })

        # Count findings by UID
        uid_counts = {}
        for f in s2_findings:
            uid_counts[f["uid"]] = uid_counts.get(f["uid"], 0) + 1

        results.append({
            "doc_name": doc_name,
            "text_length": len(text),
            "stats": stats,
            "profiles": profiles,
            "avg_profile_conf": avg_profile_conf,
            "term_stats": term_stats,
            "undefined_terms": undefined_terms,
            "defined_terms": [r.term for r in term_result.defined_terms],
            "contradictions": contra_list,
            "s2_findings": s2_findings,
            "uid_counts": uid_counts,
            "error": None,
        })

        print(
            f"         segs={stats['total_segments']}  "
            f"profiles={len(profiles)}  "
            f"s2_findings={len(s2_findings)}  "
            f"({', '.join(f'{k}:{v}' for k,v in sorted(uid_counts.items()))})"
        )

    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print(f"         ERROR: {e}")
        results.append({
            "doc_name": doc_name,
            "error": str(e),
            "traceback": err,
        })

print(f"\nProcessed {len(results)} documents. Writing Excel output...\n")

# ── write Excel ───────────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

ORANGE = "E87722"
WHITE  = "FFFFFF"

def hfont():  return Font(bold=True, color=WHITE, size=10)
def hfill():  return PatternFill("solid", fgColor=ORANGE)
def bfont():  return Font(size=9)
def waln():   return Alignment(wrap_text=True, vertical="top")
def taln():   return Alignment(vertical="top")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont()
        cell.fill = hfill()
        cell.alignment = taln()

def auto_width(ws, min_w=8, max_w=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        mx = max((len(str(cell.value)) for cell in col if cell.value), default=min_w)
        ws.column_dimensions[letter].width = min(max(mx + 2, min_w), max_w)

wb = openpyxl.Workbook()

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

ws_sum = wb.active
ws_sum.title = "Summary"

FAIL_IDS = ["FAIL-001","FAIL-002","FAIL-003","FAIL-004","FAIL-005","FAIL-006","FAIL-007","FAIL-008"]

sum_headers = [
    "Document", "Text Length", "Total Segments",
    "Normative", "Informative", "Unknown",
    "High Conf Segs", "Semantic Profiles", "Avg Profile Conf %",
    "Defined Terms", "Undefined Terms", "Term Coverage %",
    "NLI Contradictions", "Total S2 Findings",
] + FAIL_IDS + ["Error"]

ws_sum.append(sum_headers)
style_header(ws_sum, 1, len(sum_headers))

for r in results:
    if r.get("error") and "stats" not in r:
        ws_sum.append([r["doc_name"]] + [""] * (len(sum_headers) - 2) + [r["error"]])
        continue
    s = r["stats"]
    ts = r["term_stats"]
    cov = safe_float(ts.get("definition_coverage"))
    uid_counts = r.get("uid_counts", {})
    row = [
        r["doc_name"],
        r["text_length"],
        s.get("total_segments", 0),
        s.get("normative_segments", 0),
        s.get("informative_segments", 0),
        s.get("unknown_segments", 0),
        s.get("high_confidence_segments", 0),
        len(r["profiles"]),
        round(r["avg_profile_conf"] * 100, 1) if r["avg_profile_conf"] else 0,
        ts.get("defined_terms", 0),
        ts.get("undefined_terms", 0),
        round(cov * 100, 1) if cov else 0,
        len(r["contradictions"]),
        len(r.get("s2_findings", [])),
    ] + [uid_counts.get(uid, 0) for uid in FAIL_IDS] + [r.get("error") or ""]
    ws_sum.append(row)

for row in ws_sum.iter_rows(min_row=2):
    for cell in row:
        cell.font = bfont()
        cell.alignment = taln()

auto_width(ws_sum)
ws_sum.freeze_panes = "A2"

# ── Sheet 2: S2 Findings (all 8 FAIL codes) ──────────────────────────────────

ws_s2 = wb.create_sheet("S2 Findings")
s2_headers = [
    "Document", "UID", "Error Type", "Category", "Severity",
    "Confidence %", "Segment Type", "Text Span", "Explanation", "Context",
]
ws_s2.append(s2_headers)
style_header(ws_s2, 1, len(s2_headers))

SEVERITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}

for r in results:
    if r.get("error") and "s2_findings" not in r:
        continue
    findings = sorted(
        r.get("s2_findings", []),
        key=lambda f: (f["uid"], SEVERITY_ORDER.get(f["severity"], 9)),
    )
    for f in findings:
        conf = safe_float(f["confidence"])
        ws_s2.append([
            r["doc_name"],
            f["uid"],
            f["error_type"],
            f["category"],
            f["severity"],
            round(conf * 100, 1) if conf else 0,
            f["segment_type"],
            f["text_span"][:120],
            f["explanation"][:400],
            f["context"][:300],
        ])

for row in ws_s2.iter_rows(min_row=2):
    for cell in row:
        cell.font = bfont()
        cell.alignment = waln()

auto_width(ws_s2)
ws_s2.column_dimensions["I"].width = 60
ws_s2.column_dimensions["J"].width = 60
ws_s2.freeze_panes = "A2"

# ── Sheet 3: Undefined Terms (FAIL-001 detail) ────────────────────────────────

ws_terms = wb.create_sheet("Undefined Terms")
terms_headers = ["Document", "Term", "Is Acronym", "Suggested Definition", "Context Excerpt"]
ws_terms.append(terms_headers)
style_header(ws_terms, 1, len(terms_headers))

for r in results:
    if r.get("error") and "undefined_terms" not in r:
        continue
    for t in r.get("undefined_terms", []):
        ws_terms.append([
            r["doc_name"],
            t["term"],
            "Yes" if t["is_acronym"] else "No",
            t["suggested_def"] or "",
            (t["context"] or "")[:300],
        ])

for row in ws_terms.iter_rows(min_row=2):
    for cell in row:
        cell.font = bfont()
        cell.alignment = waln()

auto_width(ws_terms)
ws_terms.column_dimensions["E"].width = 60
ws_terms.freeze_panes = "A2"

# ── Sheet 4: Contradictions (FAIL-003 detail) ────────────────────────────────

ws_contra = wb.create_sheet("Contradictions")
contra_headers = ["Document", "Type", "Confidence %", "Statement A", "Statement B", "Explanation"]
ws_contra.append(contra_headers)
style_header(ws_contra, 1, len(contra_headers))

for r in results:
    if r.get("error") and "contradictions" not in r:
        continue
    for c in r.get("contradictions", []):
        conf = safe_float(c["confidence"])
        ws_contra.append([
            r["doc_name"],
            c["type"],
            round(conf * 100, 1) if conf else 0,
            c["statement_a"],
            c["statement_b"],
            c["explanation"],
        ])

for row in ws_contra.iter_rows(min_row=2):
    for cell in row:
        cell.font = bfont()
        cell.alignment = waln()

auto_width(ws_contra)
ws_contra.column_dimensions["D"].width = 55
ws_contra.column_dimensions["E"].width = 55
ws_contra.column_dimensions["F"].width = 50
ws_contra.freeze_panes = "A2"

# ── Sheet 5: Semantic Profiles ────────────────────────────────────────────────

ws_prof = wb.create_sheet("Semantic Profiles")
prof_headers = [
    "Document", "Sentence", "Subject", "Object",
    "Modality", "Condition", "Temporal", "Negation", "Confidence %",
]
ws_prof.append(prof_headers)
style_header(ws_prof, 1, len(prof_headers))

for r in results:
    if r.get("error") and "profiles" not in r:
        continue
    for p in r.get("profiles", []):
        conf = safe_float(p["confidence"])
        ws_prof.append([
            r["doc_name"],
            p["sentence"],
            p["subject"] or "",
            p["object"] or "",
            p["modality"] or "",
            p["condition"] or "",
            p["temporal"] or "",
            p["negation"] or "",
            round(conf * 100, 1) if conf else 0,
        ])

for row in ws_prof.iter_rows(min_row=2):
    for cell in row:
        cell.font = bfont()
        cell.alignment = waln()

auto_width(ws_prof)
ws_prof.column_dimensions["B"].width = 70
ws_prof.freeze_panes = "A2"

# ── save ──────────────────────────────────────────────────────────────────────

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = OUTPUT_DIR / f"pilot_results_{timestamp}.xlsx"
wb.save(out_path)

print(f"\nExcel saved: {out_path}")
print(f"\n{'='*60}")
print(f"SUMMARY")
print(f"{'='*60}")

ok = [r for r in results if not (r.get("error") and "stats" not in r)]
print(f"Documents processed:   {len(ok)} / {len(results)}")

total_s2 = sum(len(r.get("s2_findings", [])) for r in ok)
print(f"Total S2 findings:     {total_s2}")
for uid in FAIL_IDS:
    count = sum(r.get("uid_counts", {}).get(uid, 0) for r in ok)
    print(f"  {uid}: {count}")
